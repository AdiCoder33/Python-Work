import os
from pathlib import Path

from filelock import FileLock
from openpyxl import Workbook, load_workbook

from .config import DATA_DIR, TASKS_FILE, TASKS_LOCK, USERS_FILE, USERS_LOCK


USER_COLUMNS = [
    "user_id",
    "username",
    "password_hash",
    "role",
    "is_active",
    "created_at",
    "last_login_at",
]

TASK_COLUMN_DEFS = [
    {"name": "sno", "type": "int", "is_numeric": True, "is_date": False},
    {"name": "sub_division", "type": "text", "is_numeric": False, "is_date": False},
    {"name": "account_code", "type": "text", "is_numeric": False, "is_date": False},
    {"name": "number_of_works", "type": "int", "is_numeric": True, "is_date": False},
    {"name": "estimate_amount", "type": "float", "is_numeric": True, "is_date": False},
    {"name": "agreement_amount", "type": "float", "is_numeric": True, "is_date": False},
    {
        "name": "exp_upto_31_03_2025",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {
        "name": "balance_amount_as_on_01_04_2025",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {
        "name": "exp_upto_last_month",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {
        "name": "exp_during_this_month",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {
        "name": "total_exp_during_year",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {
        "name": "total_value_work_done_from_beginning",
        "type": "float",
        "is_numeric": True,
        "is_date": False,
    },
    {"name": "works_completed", "type": "int", "is_numeric": True, "is_date": False},
    {"name": "balance_works", "type": "int", "is_numeric": True, "is_date": False},
    {"name": "created_by", "type": "text", "is_numeric": False, "is_date": False},
    {"name": "created_at", "type": "date", "is_numeric": False, "is_date": True},
]

TASK_COLUMNS = [col["name"] for col in TASK_COLUMN_DEFS]
TASK_NUMERIC_COLUMNS = [col["name"] for col in TASK_COLUMN_DEFS if col["is_numeric"]]
TASK_DATE_COLUMNS = [col["name"] for col in TASK_COLUMN_DEFS if col["is_date"]]
TASK_INT_FIELDS = {col["name"] for col in TASK_COLUMN_DEFS if col["type"] == "int"}
TASK_FLOAT_FIELDS = {col["name"] for col in TASK_COLUMN_DEFS if col["type"] == "float"}
TASK_TOTAL_COLUMNS = [
    "number_of_works",
    "estimate_amount",
    "agreement_amount",
    "exp_upto_31_03_2025",
    "balance_amount_as_on_01_04_2025",
    "exp_upto_last_month",
    "exp_during_this_month",
    "total_exp_during_year",
    "total_value_work_done_from_beginning",
    "works_completed",
    "balance_works",
]


def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def safe_save_workbook(workbook, path: Path):
    tmp_path = path.with_name(path.name + ".tmp")
    with open(tmp_path, "wb") as handle:
        workbook.save(handle)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(str(tmp_path), str(path))


def ensure_users_file():
    ensure_data_dir()
    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(USER_COLUMNS)
        safe_save_workbook(wb, USERS_FILE)


def ensure_tasks_file():
    ensure_data_dir()
    if not TASKS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "tasks"
        ws.append(TASK_COLUMNS)
        safe_save_workbook(wb, TASKS_FILE)


def _normalize_user_row(row_data: dict) -> dict:
    row_data["user_id"] = str(row_data.get("user_id") or "")
    row_data["username"] = str(row_data.get("username") or "")
    row_data["password_hash"] = str(row_data.get("password_hash") or "")
    row_data["role"] = str(row_data.get("role") or "")
    row_data["is_active"] = int(row_data.get("is_active") or 0)
    row_data["created_at"] = str(row_data.get("created_at") or "")
    row_data["last_login_at"] = str(row_data.get("last_login_at") or "")
    return row_data


def _public_user_row(row_data: dict) -> dict:
    return {
        "user_id": row_data.get("user_id", ""),
        "username": row_data.get("username", ""),
        "role": row_data.get("role", ""),
        "is_active": int(row_data.get("is_active") or 0),
        "created_at": row_data.get("created_at", ""),
        "last_login_at": row_data.get("last_login_at", ""),
    }


def find_user(username: str):
    ensure_users_file()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = dict(zip(USER_COLUMNS, row))
            if row_data.get("username") == username:
                return _normalize_user_row(row_data)
    return None


def list_users(q: str | None = None, role: str | None = None, is_active: int | None = None):
    ensure_users_file()
    query = (q or "").strip().lower()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = _normalize_user_row(dict(zip(USER_COLUMNS, row)))
            username = row_data.get("username", "")
            if query and query not in username.lower():
                continue
            if role and row_data.get("role") != role:
                continue
            if is_active is not None and int(row_data.get("is_active", 0)) != is_active:
                continue
            results.append(_public_user_row(row_data))
        return results


def append_user(user_data: dict):
    ensure_users_file()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == user_data["username"]:
                raise ValueError("Username already exists.")
        ws.append([user_data.get(col, "") for col in USER_COLUMNS])
        safe_save_workbook(wb, USERS_FILE)


def update_last_login(username: str, last_login_at: str):
    ensure_users_file()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        username_col = USER_COLUMNS.index("username") + 1
        login_col = USER_COLUMNS.index("last_login_at") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=username_col).value == username:
                ws.cell(row=row, column=login_col).value = last_login_at
                safe_save_workbook(wb, USERS_FILE)
                return True
    return False


def update_user_status(username: str, is_active: int):
    ensure_users_file()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        username_col = USER_COLUMNS.index("username") + 1
        active_col = USER_COLUMNS.index("is_active") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=username_col).value == username:
                ws.cell(row=row, column=active_col).value = int(is_active)
                safe_save_workbook(wb, USERS_FILE)
                return True
    return False


def update_user_password(username: str, password_hash: str):
    ensure_users_file()
    with FileLock(str(USERS_LOCK)):
        wb = load_workbook(USERS_FILE)
        ws = wb["users"]
        username_col = USER_COLUMNS.index("username") + 1
        pass_col = USER_COLUMNS.index("password_hash") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=username_col).value == username:
                ws.cell(row=row, column=pass_col).value = password_hash
                safe_save_workbook(wb, USERS_FILE)
                return True
    return False


def _get_next_sno(ws):
    for row in range(ws.max_row, 1, -1):
        value = ws.cell(row=row, column=1).value
        if value is None:
            continue
        try:
            return int(value) + 1
        except (TypeError, ValueError):
            continue
    return 1


def append_task(task_data: dict) -> int:
    ensure_tasks_file()
    with FileLock(str(TASKS_LOCK)):
        wb = load_workbook(TASKS_FILE)
        ws = wb["tasks"]
        sno = _get_next_sno(ws)
        task_row = dict(task_data)
        task_row["sno"] = sno
        ws.append([task_row.get(col, "") for col in TASK_COLUMNS])
        safe_save_workbook(wb, TASKS_FILE)
        return sno


def _to_int(value):
    if value is None or value == "":
        return 0
    return int(value)


def _to_float(value):
    if value is None or value == "":
        return 0.0
    return float(value)


def _normalize_task_row(row_data: dict) -> dict:
    normalized = {}
    for col in TASK_COLUMN_DEFS:
        key = col["name"]
        value = row_data.get(key)
        if col["type"] == "int":
            normalized[key] = _to_int(value)
        elif col["type"] == "float":
            normalized[key] = _to_float(value)
        else:
            normalized[key] = str(value or "")
    return normalized


def list_tasks():
    ensure_tasks_file()
    with FileLock(str(TASKS_LOCK)):
        wb = load_workbook(TASKS_FILE)
        ws = wb["tasks"]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = dict(zip(TASK_COLUMNS, row))
            records.append(_normalize_task_row(row_data))
        return records


def copy_tasks_backup(backup_path: Path | None = None):
    ensure_tasks_file()
    if backup_path is None:
        backup_path = DATA_DIR / "tasks_backup.xlsx"
    with FileLock(str(TASKS_LOCK)):
        wb = load_workbook(TASKS_FILE)
        safe_save_workbook(wb, backup_path)
