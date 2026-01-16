from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from fastapi import Depends, FastAPI, HTTPException, status
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import FileResponse
from jose import JWTError, jwt
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, validator
from filelock import FileLock

SECRET_KEY = "change-this-secret"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

# Simple in-memory user store for demo purposes
USERS_DB: Dict[str, Dict[str, str]] = {
    "admin": {"username": "admin", "password": "admin123", "role": "Admin"},
    "user": {"username": "user", "password": "user123", "role": "User"},
}

app = FastAPI(title="Work Progress Backend", version="1.0.0")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/token")

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / "work_entries.xlsx"
LOCK_PATH = EXCEL_PATH.with_suffix(".lock")
FIELD_ORDER = [
    "s_no",
    "sub_division",
    "account_code",
    "num_works",
    "estimate_amount",
    "agreement_amount",
    "exp_3103",
    "balance_0104",
    "exp_last_month",
    "exp_this_month",
    "total_year",
    "total_from_beginning",
    "completed_works",
    "balance_works",
]
HEADERS = [
    "S. No.",
    "Sub-Division",
    "Account Code",
    "Number of Works",
    "Estimate Amount",
    "Agreement Amount",
    "Expenditure up to 31-03-2025",
    "Balance Amount as on 01-04-2025",
    "Expenditure up to Last Month",
    "Expenditure During This Month",
    "Total Expenditure During the Year",
    "Total Value of Work Done from the Beginning",
    "Number of Works Completed",
    "Balance Number of Works",
]


class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"
    role: str


class TokenData(BaseModel):
    username: str
    role: str


class FormPayload(BaseModel):
    s_no: int = Field(..., ge=1)
    sub_division: str
    account_code: str
    num_works: int = Field(..., ge=1)
    estimate_amount: float = Field(..., ge=0)
    agreement_amount: float = Field(..., ge=0)
    exp_3103: float = Field(..., ge=0)
    balance_0104: float = Field(..., ge=0)
    exp_last_month: float = Field(..., ge=0)
    exp_this_month: float = Field(..., ge=0)
    total_year: float = Field(..., ge=0)
    total_from_beginning: float = Field(..., ge=0)
    completed_works: int = Field(..., ge=0)
    balance_works: int = Field(..., ge=0)

    @validator("account_code")
    def validate_account_code(cls, value: str) -> str:
        if value not in {"Spill", "New"}:
            raise ValueError("Account Code must be Spill or New")
        return value

    @validator("sub_division")
    def validate_sub_division(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("Sub-Division is required")
        return value.strip()

    @validator("completed_works")
    def validate_completed_vs_total(cls, value: int, values: Dict[str, object]) -> int:
        total = values.get("num_works", 0)
        if value < 0 or value > total:
            raise ValueError("Completed works cannot exceed total works")
        return value

    def recalc(self) -> Dict[str, object]:
        balance_start = round(self.estimate_amount - self.exp_3103, 2)
        total_year = round(self.exp_last_month + self.exp_this_month, 2)
        total_from_beginning = round(self.exp_3103 + total_year, 2)
        balance_works = self.num_works - self.completed_works

        return {
            "s_no": self.s_no,
            "sub_division": self.sub_division.strip(),
            "account_code": self.account_code,
            "num_works": self.num_works,
            "estimate_amount": round(self.estimate_amount, 2),
            "agreement_amount": round(self.agreement_amount, 2),
            "exp_3103": round(self.exp_3103, 2),
            "balance_0104": balance_start,
            "exp_last_month": round(self.exp_last_month, 2),
            "exp_this_month": round(self.exp_this_month, 2),
            "total_year": total_year,
            "total_from_beginning": total_from_beginning,
            "completed_works": self.completed_works,
            "balance_works": balance_works,
        }


class User(BaseModel):
    username: str
    role: str


def verify_user(username: str, password: str) -> Optional[User]:
    user = USERS_DB.get(username)
    if user and password == user.get("password"):
        return User(username=user["username"], role=user["role"])
    return None


def create_access_token(data: Dict[str, str], expires_delta: timedelta) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + expires_delta
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme)) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None or role is None:
            raise credentials_exception
        return User(username=username, role=role)
    except JWTError:
        raise credentials_exception


def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != "Admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")
    return current_user


@app.post("/token", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = verify_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect username or password")

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "role": user.role}, expires_delta=access_token_expires
    )
    return Token(access_token=access_token, token_type="bearer", role=user.role)


@app.post("/submit")
def submit_form(payload: FormPayload, current_user: User = Depends(get_current_user)):
    recalculated = payload.recalc()
    append_row_to_excel(recalculated)
    return {"message": "Record stored", "record": recalculated, "submitted_by": current_user.username}


@app.get("/records")
def list_records(current_user: User = Depends(require_admin)):
    records = read_excel_records()
    return {"records": records}


@app.get("/export")
def export_excel(current_user: User = Depends(require_admin)):
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail="No data file found")
    return FileResponse(EXCEL_PATH, filename="work_entries.xlsx", media_type="application/vnd.ms-excel")


def append_row_to_excel(record: Dict[str, object]) -> None:
    lock = FileLock(LOCK_PATH)
    with lock:
        if not EXCEL_PATH.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            wb.save(EXCEL_PATH)

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([record.get(key, "") for key in FIELD_ORDER])
        wb.save(EXCEL_PATH)


def read_excel_records() -> List[Dict[str, object]]:
    numeric_fields = {
        "s_no",
        "num_works",
        "estimate_amount",
        "agreement_amount",
        "exp_3103",
        "balance_0104",
        "exp_last_month",
        "exp_this_month",
        "total_year",
        "total_from_beginning",
        "completed_works",
        "balance_works",
    }

    if not EXCEL_PATH.exists():
        return []

    lock = FileLock(LOCK_PATH)
    with lock:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        records: List[Dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ensure row length matches the field order; pad if needed.
            padded = list(row) + [None] * (len(FIELD_ORDER) - len(row))
            record: Dict[str, object] = {}
            for idx, key in enumerate(FIELD_ORDER):
                value = padded[idx]
                if value is None:
                    record[key] = 0 if key in numeric_fields else ""
                    continue
                if key in numeric_fields:
                    # Coerce numeric fields to float or int where appropriate.
                    if key in {"s_no", "num_works", "completed_works", "balance_works"}:
                        record[key] = int(value)
                    else:
                        record[key] = float(value)
                else:
                    record[key] = str(value)
            records.append(record)
        return records


@app.get("/health")
def health_check():
    return {"status": "ok"}
